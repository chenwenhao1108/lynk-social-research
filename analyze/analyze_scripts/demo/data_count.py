"""
用于统计分析后的数据的酒店总buzz，一级关键词和二级关键词的情感分布和情感得分，运行该脚本会在analysis_result文件夹下生成data_count.xlsx文件
"""

from openpyxl import Workbook
from utils import *
import pandas as pd
import json  # Ensure json is imported
from openpyxl import Workbook  # For creating new Excel files
from openpyxl.utils.dataframe import (
    dataframe_to_rows,
)  # For writing dataframes to sheets


def caculate_sentiment_distribution(sentiment_distribution):
    if sentiment_distribution is None:
        sentiment_distribution = {}

    positive = sentiment_distribution.get("positive", 0)
    negative = sentiment_distribution.get("negative", 0)
    neutral = sentiment_distribution.get("neutral", 0)

    total_buzz = positive + negative + neutral
    raw_score = positive - negative

    sentiment_score_percent = (raw_score / total_buzz) * 100 if total_buzz > 0 else 0

    return {"totalBuzz": total_buzz, "sentimentScorePercent": sentiment_score_percent}


def compile_keywords_for_analyzed_data(analyzed_data):
    """
    统计每个酒店每个关键词的情感分布以及每个酒店的buzz数量
    """
    sk_to_pk_map = Keywords.get_sk_to_pk_map()
    valid_sentiments = ["positive", "negative", "neutral"]
    compiled_data = {}

    initialSentimentDistribution = {}
    for sk, pk in sk_to_pk_map.items():
        initialSentimentDistribution[sk] = {
            "positive": 0,
            "negative": 0,
            "neutral": 0,
        }
        if not initialSentimentDistribution.get(pk):
            initialSentimentDistribution[pk] = {
                "positive": 0,
                "negative": 0,
                "neutral": 0,
            }

    for hotel_entry in analyzed_data:
        hotel_name = hotel_entry["hotel"]
        if not compiled_data.get(hotel_name):
            compiled_data[hotel_name] = {
                "buzz": 0,
                "keywords_sentiment_distribution": json.loads(
                    json.dumps(initialSentimentDistribution)
                ),
            }
        for post in hotel_entry["posts"]:
            if not post.get("is_hotel_related", False):
                continue
            compiled_data[hotel_name]["buzz"] += 1

            replies = post.get("replies", [])
            replies_count = len(replies)
            compiled_data[hotel_name]["buzz"] += replies_count

            # 处理帖子本身
            primary_keywords = post.get("keywords_mentioned", {}).get(
                "primary_keyword", []
            )
            secondary_keywords = post.get("keywords_mentioned", {}).get(
                "secondary_keyword", []
            )
            for keyword_dict in primary_keywords + secondary_keywords:
                keyword = keyword_dict.get("keyword", "").strip()
                sentiment = keyword_dict.get("sentiment", "")
                if not keyword or sentiment not in valid_sentiments:
                    continue
                compiled_data[hotel_name]["keywords_sentiment_distribution"][keyword][
                    sentiment
                ] += (1 + replies_count)
                # 如果是二级关键词，则对应的一级关键词也要加上
                if not Keywords.is_primary_keyword(keyword):
                    pk = sk_to_pk_map[keyword]
                    compiled_data[hotel_name]["keywords_sentiment_distribution"][pk][
                        sentiment
                    ] += (1 + replies_count)

            # 处理帖子回复
            for reply in replies:
                reply_keywords = reply.get("keywords_mentioned", {}).get(
                    "primary_keyword", []
                ) + reply.get("keywords_mentioned", {}).get("secondary_keyword", [])
                for keyword_dict in reply_keywords:
                    keyword = keyword_dict.get("keyword", "").strip()
                    sentiment = keyword_dict.get("sentiment", "")
                    if not keyword or sentiment not in valid_sentiments:
                        continue
                    compiled_data[hotel_name]["keywords_sentiment_distribution"][
                        keyword
                    ][sentiment] += 1
                    # 如果是二级关键词，则对应的一级关键词也要加上
                    if not Keywords.is_primary_keyword(keyword):
                        pk = sk_to_pk_map[keyword]
                        compiled_data[hotel_name]["keywords_sentiment_distribution"][
                            pk
                        ][sentiment] += 1
    return compiled_data


def generate_excel_for_compiled_data(compiled_data, excel_file_path):
    """
    生成excel表格
    """
    # 创建Excel工作簿
    wb = Workbook()
    # 删除默认创建的sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    column_order = [
        "酒店声量",
        "一级关键词",
        "一级正面",
        "一级中立",
        "一级负面",
        "一级关键词情感得分",
        "二级关键词",
        "二级正面",
        "二级中立",
        "二级负面",
        "二级关键词情感得分",
    ]

    sk_to_pk_map = Keywords.get_sk_to_pk_map()

    for hotel_name, hotel_data in compiled_data.items():
        pk_data = []
        sk_data = {}
        hotel_buzz = hotel_data.get("buzz", 0)
        keywords_sentiment_distribution = hotel_data.get(
            "keywords_sentiment_distribution", {}
        )
        # 收集一级关键词和二级关键词
        for keyword, sentiment_distribution in keywords_sentiment_distribution.items():
            if Keywords.is_primary_keyword(keyword):
                pk_data.append(
                    {
                        "keyword": keyword,
                        "sentiment_distribution": sentiment_distribution,
                    }
                )
            else:
                pk = sk_to_pk_map.get(keyword)
                if not sk_data.get(pk):
                    sk_data[pk] = []
                sk_data[pk].append(
                    {
                        "keyword": keyword,
                        "sentiment_distribution": sentiment_distribution,
                    }
                )

        hotel_specific_rows = []  # 用于存储当前酒店的行数据
        for pk_dict in sorted(pk_data, key=lambda x: x["keyword"]):
            pk_name = pk_dict["keyword"]
            pk_sentiment = pk_dict["sentiment_distribution"]
            pk_score = caculate_sentiment_distribution(pk_sentiment).get(
                "sentimentScorePercent", 0
            )
            row_data_pk = {
                "酒店声量": hotel_buzz,
                "一级关键词": pk_name,
                "一级正面": pk_sentiment.get("positive", 0),
                "一级中立": pk_sentiment.get("neutral", 0),
                "一级负面": pk_sentiment.get("negative", 0),
                "一级关键词情感得分": f"{pk_score:.2f}%",
                "二级关键词": "",
                "二级正面": "",
                "二级中立": "",
                "二级负面": "",
                "二级关键词情感得分": "",
            }
            hotel_specific_rows.append(row_data_pk)

            for sk_dict in sorted(sk_data.get(pk_name, []), key=lambda x: x["keyword"]):
                sk_name = sk_dict["keyword"]
                sk_sentiment = sk_dict["sentiment_distribution"]
                sk_score = caculate_sentiment_distribution(sk_sentiment).get(
                    "sentimentScorePercent", 0
                )
                row_data_sk = {
                    "酒店声量": hotel_buzz,
                    "一级关键词": pk_name,
                    "一级正面": pk_sentiment.get("positive", 0),
                    "一级中立": pk_sentiment.get("neutral", 0),
                    "一级负面": pk_sentiment.get("negative", 0),
                    "一级关键词情感得分": f"{pk_score:.2f}%",
                    "二级关键词": sk_name,
                    "二级正面": sk_sentiment.get("positive", 0),
                    "二级中立": sk_sentiment.get("neutral", 0),
                    "二级负面": sk_sentiment.get("negative", 0),
                    "二级关键词情感得分": f"{sk_score:.2f}%",
                }
                hotel_specific_rows.append(row_data_sk)

        if not hotel_specific_rows:  # 如果当前酒店没有数据
            # 创建一个带表头的空sheet
            df = pd.DataFrame(columns=column_order)  # type: ignore
        else:
            df = pd.DataFrame(hotel_specific_rows)
            # 检查df是否为空，以及是否包含所有column_order中的列
            if not df.empty and all(col in df.columns for col in column_order):
                df = df[column_order]

        # 创建新的sheet
        ws = wb.create_sheet(title=hotel_name)

        # 写入表头
        ws.append(column_order)

        # 写入数据
        for r_idx, row in enumerate(
            dataframe_to_rows(df, index=False, header=False), 2
        ):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save(excel_file_path)


def get_all_analyzed_data(file_paths):
    """
    读取所有的analyzed_data
    """
    all_analyzed_data = []
    for file_path in file_paths:
        data = get_raw_data(file_path)
        if data:
            all_analyzed_data.extend(data)
    return all_analyzed_data


if __name__ == "__main__":
    file_paths = [
        "analysis_result/flyert_analyzed.json",
        "analysis_result/wb_analyzed.json",
        "analysis_result/xhs_analyzed.json",
    ]
    excel_file_path = "analysis_result/data_count.xlsx"
    analyzed_data = get_all_analyzed_data(file_paths)
    compiled_data = compile_keywords_for_analyzed_data(analyzed_data)
    generate_excel_for_compiled_data(compiled_data, excel_file_path)
    print(f"统计结果已保存在{excel_file_path}")
